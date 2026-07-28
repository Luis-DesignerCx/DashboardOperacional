"""
Extrai linhas FP/PON do Excel e salva como JSON.
Uso: python scripts/fapass-extrair.py "caminho/Base CAR Passaporte BC.xlsx"
Saída: scripts/fapass-dados.json
"""
import sys, json, re
from openpyxl import load_workbook

PREFIXOS = ("FP", "PON")

def parsear_data(val):
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}T00:00:00Z"
    # Pode vir como datetime do openpyxl
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%dT00:00:00Z")
    return None

def parsear_valor(val):
    if val is None:
        return 0
    try:
        return abs(float(str(val).replace(".", "").replace(",", ".")))
    except:
        return 0

arquivo = sys.argv[1] if len(sys.argv) > 1 else "planilhas/Base CAR Passaporte BC.xlsx"
saida = "scripts/fapass-dados.json"

print(f"Abrindo {arquivo} em modo streaming...")
wb = load_workbook(filename=arquivo, read_only=True, data_only=True)
ws = wb.active

linhas_fp = []
cabecalho = None
total = 0
encontradas = 0

for row in ws.iter_rows(values_only=True):
    if cabecalho is None:
        cabecalho = list(row)
        print(f"Cabeçalho ({len(cabecalho)} colunas): {cabecalho[:10]}...")
        continue

    total += 1
    if total % 50000 == 0:
        print(f"  Lidas {total} linhas, {encontradas} FP/PON encontradas...")

    # Coluna B = índice 1
    doc = str(row[1] or "").strip().upper()
    if not any(doc.startswith(p) for p in PREFIXOS):
        continue

    encontradas += 1
    linhas_fp.append({
        "documento":  str(row[1]  or "").strip().upper(),
        "tipo":       str(row[6]  or "").strip(),
        "fornecedor": str(row[7]  or "").strip(),
        "valor":      parsear_valor(row[8]),
        "saldo":      parsear_valor(row[14]),
        "status":     str(row[16] or "").strip().upper(),
        "vencimento": parsear_data(row[17]),
        "dataBaixa":  parsear_data(row[18]),
        "tiposBaixa": str(row[19] or "").strip(),
    })

wb.close()
print(f"\nTotal lidas: {total} | FP/PON encontradas: {encontradas}")

with open(saida, "w", encoding="utf-8") as f:
    json.dump(linhas_fp, f, ensure_ascii=False)

print(f"Dados salvos em {saida} ({len(linhas_fp)} registros)")
