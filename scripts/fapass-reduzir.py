"""
Reduz uma planilha bruta do Fã Pass ("Base CAR Passaporte BC") mantendo só
as linhas com Documento prefixado por FP ou PON — as únicas que a
importação (src/app/api/fapass/importar/route.ts) realmente usa. Mantém
todas as colunas na mesma posição original, então não quebra o
mapeamento de colunas.

Útil quando o arquivo bruto passa do limite de 50MB do bucket do Supabase
Storage (plano gratuito) — a maior parte do arquivo é dado histórico já
baixado/irrelevante, então filtrar por prefixo costuma reduzir bastante
o tamanho sem perder nenhuma linha que a importação usaria de qualquer
forma.

Uso: python scripts/fapass-reduzir.py "caminho/Base CAR Passaporte BC.xlsx"
Saída: mesmo caminho + " - reduzida.xlsx"
"""
import sys
import os
from openpyxl import load_workbook, Workbook

origem = sys.argv[1] if len(sys.argv) > 1 else "planilhas/Base CAR Passaporte BC.xlsx"
nome, ext = os.path.splitext(origem)
destino = f"{nome} - reduzida{ext}"

print(f"Lendo {origem} em modo streaming...")
wb_in = load_workbook(filename=origem, read_only=True, data_only=True)
ws_in = wb_in.active

wb_out = Workbook(write_only=True)
ws_out = wb_out.create_sheet(ws_in.title or "Base CAR Passaporte")

total = 0
mantidas = 0

for i, row in enumerate(ws_in.iter_rows(values_only=True)):
    if i == 0:
        ws_out.append(list(row))
        continue
    total += 1
    doc = str(row[1] or "").strip().upper()
    if doc.startswith("FP") or doc.startswith("PON"):
        ws_out.append(list(row))
        mantidas += 1
    if total % 50000 == 0:
        print(f"  processadas {total}, mantidas {mantidas}...")

wb_in.close()
wb_out.save(destino)

tam_mb = os.path.getsize(destino) / 1_048_576

print(f"\nTotal linhas originais: {total}")
print(f"Linhas mantidas (prefixo FP/PON): {mantidas}")
print(f"Arquivo salvo em: {destino}")
print(f"Tamanho final: {tam_mb:.2f} MB")
if tam_mb > 50:
    print("\nAVISO: ainda está acima de 50MB (limite do bucket do Supabase Storage).")
